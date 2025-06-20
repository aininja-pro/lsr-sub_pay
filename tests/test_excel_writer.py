import pandas as pd
import pytest
import os
import io
import tempfile
import openpyxl
from datetime import datetime, timedelta
from src.utils.excel_writer import create_pay_sheet

class MockFileUpload:
    """Mock class to simulate a file upload in Streamlit."""
    def __init__(self, content):
        self.content = content
    
    def getvalue(self):
        return self.content

def create_test_template():
    """Create a simple test template workbook."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets for test subcontractors
    for sub in ["Sub 1", "Sub 2"]:
        sheet = wb.create_sheet(title=sub)
        
        # Create header row at row 12
        sheet.cell(row=12, column=1).value = "Date"
        sheet.cell(row=12, column=2).value = "Property"
        sheet.cell(row=12, column=3).value = "Job #"
        sheet.cell(row=12, column=4).value = "Description"
        sheet.cell(row=12, column=5).value = "Qty"
        sheet.cell(row=12, column=6).value = "Per Unit"
        sheet.cell(row=12, column=7).value = "Amount"
        
        # Create summary row at row 30
        sheet.cell(row=30, column=4).value = "Total"
        sheet.cell(row=30, column=7).value = "=SUM(G13:G29)"
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

def test_create_pay_sheet():
    """Test creating a pay sheet from filtered data."""
    # Create test data
    today = datetime.now().date()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    
    # Create DataFrame of jobs
    data = {
        'Tech': ['Sub 1', 'Sub 1', 'Sub 2', 'Sub 3'],  # Sub 3 is not in template
        'Job#': ['1001', '1002', '1003', '1004'],
        'Completed On': [monday, monday + timedelta(days=2), sunday, sunday],
        'Job Category': ['Category 1', 'Category 2', 'Category 3', 'Category 4'],
        'Missing Date': [False, False, False, False]
    }
    
    df = pd.DataFrame(data)
    
    # Create template and mock file upload
    template_content = create_test_template()
    template_file = MockFileUpload(template_content)
    
    # Call the function
    output_path, skipped_subs = create_pay_sheet(template_file, df, [monday, sunday])
    
    try:
        # Verify output
        assert os.path.exists(output_path)
        assert output_path.endswith('.xlsx')
        assert "Sub 3" in skipped_subs  # Should skip Sub 3
        
        # Load the workbook and verify content
        wb = openpyxl.load_workbook(output_path)
        
        # Check Sub 1 sheet
        sheet = wb["Sub 1"]
        assert sheet.cell(row=13, column=1).value == monday  # Date
        assert sheet.cell(row=13, column=3).value == "1001"  # Job#
        assert sheet.cell(row=14, column=3).value == "1002"  # Job#
        assert sheet.cell(row=13, column=5).value == 1  # Qty
        
        # Check Sub 2 sheet
        sheet = wb["Sub 2"]
        assert sheet.cell(row=13, column=1).value == sunday  # Date
        assert sheet.cell(row=13, column=3).value == "1003"  # Job#
        assert sheet.cell(row=13, column=5).value == 1  # Qty
        
        # Check formulas preserved
        sheet = wb["Sub 1"]
        assert sheet.cell(row=30, column=7).value == "=SUM(G13:G29)"  # Formula preserved
    
    finally:
        # Clean up
        if os.path.exists(output_path):
            os.remove(output_path)

if __name__ == "__main__":
    pytest.main(['-v', __file__]) 