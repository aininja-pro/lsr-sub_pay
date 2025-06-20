import pandas as pd
import openpyxl
import os
import tempfile
from datetime import datetime
import logging
from pathlib import Path
import dateutil.parser

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_pay_sheet(template_file, filtered_df, date_range):
    """
    Create a pay sheet from the template and filtered job data.
    
    Args:
        template_file: The uploaded template file object
        filtered_df (pandas.DataFrame): DataFrame of filtered jobs
        date_range (list): [start_date, end_date] as datetime.date objects
    
    Returns:
        tuple: (output_path, skipped_subs) - Path to the generated Excel file and list of skipped subcontractors
    """
    skipped_subs = []
    
    try:
        # Check if we have data to process
        if filtered_df.empty:
            raise ValueError("No jobs to include in the pay sheet")
        
        # Create a temporary directory to save the file
        temp_dir = tempfile.mkdtemp()
        
        # Generate output filename based on date range
        if date_range and len(date_range) == 2:
            start_date_str = date_range[0].strftime("%Y-%m-%d")
            end_date_str = date_range[1].strftime("%Y-%m-%d")
            output_filename = f"Sub_PaySheet_{start_date_str}_to_{end_date_str}.xlsx"
        else:
            # Fallback if date range is not provided
            output_filename = f"Sub_PaySheet_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        output_path = os.path.join(temp_dir, output_filename)
        
        # Save the template file to disk temporarily
        temp_template = os.path.join(temp_dir, "template.xlsx")
        with open(temp_template, "wb") as f:
            f.write(template_file.getvalue())
        
        # Load the workbook with openpyxl (preserving formulas)
        workbook = openpyxl.load_workbook(temp_template, keep_vba=False)
        
        # Get list of sheet names (case-insensitive for comparison)
        sheet_names = {name.lower().strip(): name for name in workbook.sheetnames}
        
        # Get unique subcontractors in the filtered data
        unique_subs = filtered_df['Tech'].unique()
        
        # Process each subcontractor
        for sub in unique_subs:
            # Find matching sheet in template (case-insensitive)
            sub_key = sub.lower().strip()
            
            if sub_key not in sheet_names:
                logger.warning(f"No matching sheet found for subcontractor: {sub}")
                skipped_subs.append(sub)
                continue
            
            # Get actual sheet name with correct case
            actual_sheet_name = sheet_names[sub_key]
            sheet = workbook[actual_sheet_name]
            
            # Add Week Of date range to the template (typically cell B4)
            if date_range and len(date_range) == 2:
                # Format as MM/DD/YY - MM/DD/YY
                start_date, end_date = date_range
                week_of_text = f"{start_date.strftime('%m/%d/%y')} - {end_date.strftime('%m/%d/%y')}"
                
                # Find the "Week Of:" cell and populate the cell to its right
                week_of_cell_found = False
                # Search first few rows for "Week Of:" label
                for row_idx in range(1, 10):  # Search rows 1-9
                    for col_idx in range(1, 5):  # Search first few columns
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value and "Week Of" in str(cell_value):
                            # Found the Week Of label, now populate the cell to its right
                            sheet.cell(row=row_idx, column=col_idx+1).value = week_of_text
                            logger.info(f"Added Week Of: {week_of_text} to row {row_idx}, column {col_idx+1}")
                            week_of_cell_found = True
                            break
                    if week_of_cell_found:
                        break
                
                # If we didn't find the cell, try a common location (B4)
                if not week_of_cell_found:
                    sheet.cell(row=4, column=2).value = week_of_text
                    logger.info(f"Added Week Of: {week_of_text} to default location (B4)")
            
            # Get jobs for this subcontractor
            sub_jobs = filtered_df[filtered_df['Tech'] == sub].copy()
            logger.info(f"Found {len(sub_jobs)} jobs for {sub}")
            
            # Convert to list of dictionaries for simpler sorting
            jobs_list = sub_jobs.to_dict(orient='records')
            
            # Parse dates for each job
            for job in jobs_list:
                # Extract date as a string in MM/DD/YY format
                if 'Completed On' in job and pd.notna(job['Completed On']):
                    date_str = str(job['Completed On'])
                    # Try to extract a sortable date string
                    try:
                        if isinstance(job['Completed On'], (pd.Timestamp, datetime)):
                            # Already a datetime, extract month/day/year as numbers
                            job['_month'] = job['Completed On'].month
                            job['_day'] = job['Completed On'].day
                            job['_year'] = job['Completed On'].year
                        else:
                            # Parse from string
                            date_obj = dateutil.parser.parse(date_str)
                            job['_month'] = date_obj.month
                            job['_day'] = date_obj.day
                            job['_year'] = date_obj.year
                    except:
                        # If parsing fails, set to high values to sort to end
                        job['_month'] = 99
                        job['_day'] = 99
                        job['_year'] = 9999
                else:
                    # No date, sort to end
                    job['_month'] = 99
                    job['_day'] = 99
                    job['_year'] = 9999
            
            # Simple manual sort by year, month, day
            sorted_jobs = sorted(jobs_list, key=lambda x: (x['_year'], x['_month'], x['_day']))
            
            # Debug - print the sorted dates
            debug_dates = []
            for job in sorted_jobs:
                if '_month' in job:
                    debug_dates.append(f"{job['_month']:02d}/{job['_day']:02d}/{job['_year']}")
                else:
                    debug_dates.append("No date")
            
            logger.info(f"Sorted dates for {sub}: {debug_dates}")
            
            # Set header row (row 12) and data start row (row 13)
            header_row = 12
            data_start_row = 13
            data_end_row = 29  # Maximum rows before summary row at row 30
            
            # Validate header row
            if sheet.cell(row=header_row, column=1).value is None:
                logger.warning(f"Header row (row {header_row}) in sheet '{actual_sheet_name}' appears to be empty")
            
            # Check if we'll exceed the available rows
            max_rows = min(len(sorted_jobs), data_end_row - data_start_row + 1)
            if len(sorted_jobs) > max_rows:
                logger.warning(f"Only {max_rows} of {len(sorted_jobs)} jobs will be included for {sub} due to template limits")
            
            # Add jobs to the sheet, using the sorted list
            for i, job in enumerate(sorted_jobs[:max_rows]):
                row = data_start_row + i
                
                # Debug this row
                logger.info(f"Writing job {i+1} to row {row}: Job# {job.get('Job#', 'N/A')}, Date: {job.get('Completed On', 'N/A')}")
                
                # Date (column A)
                # Debug the date to see what we're getting
                logger.info(f"Job {job.get('Job#', 'N/A')} date value: {job.get('Completed On')}, type: {type(job.get('Completed On'))}")
                
                # Use our already extracted and parsed date components
                if '_year' in job and job['_year'] != 9999:
                    try:
                        # Create a date object from our parsed components
                        date_obj = datetime(job['_year'], job['_month'], job['_day'])
                        sheet.cell(row=row, column=1).value = date_obj.date()
                        logger.info(f"Date set for row {row}: {date_obj.date()}")
                    except Exception as e:
                        logger.warning(f"Error creating date for row {row}: {str(e)}")
                        # Fallback to direct value
                        sheet.cell(row=row, column=1).value = job.get('Completed On')
                else:
                    # No valid date
                    sheet.cell(row=row, column=1).value = None
                
                # Property/Building Unit (column B) - Use Service Location Address 1 if available
                if 'Service Location Address 1' in job and pd.notna(job['Service Location Address 1']):
                    property_value = str(job['Service Location Address 1'])
                else:
                    # Fallback to Job Category if Service Location Address 1 is not available
                    property_value = str(job['Job Category']) if 'Job Category' in job else "N/A"
                sheet.cell(row=row, column=2).value = property_value
                
                # Job Number (column C)
                if 'Job#' in job and pd.notna(job['Job#']):
                    try:
                        # Try to convert to float first, then to int to remove decimal
                        job_number = int(float(job['Job#']))
                        # Assign as a number, not a string
                        sheet.cell(row=row, column=3).value = job_number
                    except (ValueError, TypeError):
                        # If conversion fails, fallback to string
                        job_number = str(job['Job#'])
                        sheet.cell(row=row, column=3).value = job_number
                else:
                    sheet.cell(row=row, column=3).value = "N/A"
                
                # Description (column D) - Use truncated Job Details if available
                if 'Job Details' in job and pd.notna(job['Job Details']):
                    # Truncate Job Details to a reasonable length (e.g., 100 characters)
                    job_details = str(job['Job Details'])
                    description = job_details[:100] + "..." if len(job_details) > 100 else job_details
                else:
                    # Fallback to Job Category if Job Details is not available
                    description = str(job['Job Category']) if 'Job Category' in job else "N/A"
                sheet.cell(row=row, column=4).value = description
                
                # Quantity (column E) - Set to 1
                sheet.cell(row=row, column=5).value = 1
                
                # Per Unit (column F) - Leave blank for manual entry
                sheet.cell(row=row, column=6).value = None
                
                # Amount (column G) - Has formula, leave untouched
            
            logger.info(f"Added {min(len(sorted_jobs), max_rows)} jobs for {sub} to sheet '{actual_sheet_name}'")
        
        # Save the workbook
        workbook.save(output_path)
        logger.info(f"Pay sheet saved to {output_path}")
        
        return output_path, skipped_subs
    
    except Exception as e:
        logger.error(f"Error creating pay sheet: {str(e)}")
        raise 